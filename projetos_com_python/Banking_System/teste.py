from pathlib import Path
import openpyxl
import pandas as pd
import os


ROOT_FOLDER = Path(__file__).parent
WOORKBOOK = ROOT_FOLDER / 'Senhas_client.xlsx'

"""
# Carregar o arquivo Excel
arquivo = "C:\\Users\\freds\\Desktop\\curso_de_programacao\\aprender_python\\Projetos_com_python\\Banking_System\\Banco_de_Dados.xlsx"
# Carregar o arquivo Excel
workbook = openpyxl.load_workbook(arquivo)

# Selecionar a planilha desejada
sheet_selecionada = workbook.active"""



# Definir a função de comparação
""" def cliente_corresponde(nome, nif):
    return nome == 'Fredson Vila Nova' and int(nif) == 308806662

# Percorrer cada linha na planilha, começando na linha 2 para ignorar o cabeçalho
nome_encontrado = False
for row in range(2, sheet_selecionada.max_row + 1):
    name_cell = sheet_selecionada.cell(row=row, column=1).value  # Coluna "A" para os nomes
    nif_cell = sheet_selecionada.cell(row=row, column=8).value   # Coluna "H" para os NIFs

    if cliente_corresponde(name_cell, nif_cell):
        nome_encontrado = True
        break

if nome_encontrado:
    print("O nome 'Fredson Vila Nova' está presente no banco de dados.")
else:
    print("O nome 'Fredson Vila Nova' não está presente no banco de dados.") """

""" def regitar_client(arquivo_excel):
        name_clint = input('Crie um nome: ')
        senha = input('Crie uma senha: ')

        # Carregar o arquivo Excel
        workbook = openpyxl.load_workbook(arquivo_excel)

        # Verificar se a panilha "senhas_client" existe, caso contrario, criar uma nova
    
        if 'senhas_clint' not in workbook.sheetnames:
            workbook.create_sheet('senhas_client')

        # seleclionar a planilha desejada
        sheet_selecionada = workbook['senhas_client']

        # Se a planilha estiver vazia, criar um cabecalho
        if sheet_selecionada.max_row == 1:
            sheet_selecionada.cell(row=1, column=1, value = "Nome do Cliente")
            sheet_selecionada.cell(row=1, column=2, value = "Senha")

        # Adicionar o nome e a senha à proxima linha vazia
        next_row = sheet_selecionada.max_row + 1 # proxima linha vazia para inserir o nome e a senha
        sheet_selecionada.cell(row=next_row, column=1, value=name_clint)
        sheet_selecionada.cell(row=next_row, column=2, value=senha)

        # Salvar o arquivo Excel
        workbook.save(arquivo_excel)
        workbook.close()
    
    

 """

def registar_client(WOORKBOOK):
    name_client = input('Degiti o nome do cliente: ')
    senha = input('Crie uma senha: ')

    # Carregar o arquivo Excel

    
            # Verificar se a panilha "senhas_client" existe, caso contrario, criar uma nova
    try:
        #workbook = openpyxl.load_workbook(arquivo_excel)


        
        # Verificar se a panilha "senhas_client" existe, caso contrario, criar uma nova
    
        if os.path.exists(WOORKBOOK):
            if 'senhas_client' in pd.ExcelFile(WOORKBOOK).sheet_names:
                df = pd.read_excel(WOORKBOOK, sheet_name='senhas_client')
            else:
                df = pd.DataFrame(columns=['Nome do Cliente', 'Senha'])
        

    except FileNotFoundError:
        df = pd.DataFrame(columns=['Nome do Cliente', 'Senha'])
        print('Arquivo não encontrado -> registar_client')
    

    novo_Cliente = {
            'Nome do Cliente': name_client,
            'Senha': senha
            
        }

    client_df = pd.DataFrame.from_dict([novo_Cliente])

    dt = pd.concat([client_df], ignore_index=True)

    
    dt.to_excel(sheet_name='senhas_client')

    print( 'Senha guardada com Segurança -> registar_client')    



#arquivo = 'C:\\Users\\freds\\Desktop\\curso_de_programacao\\aprender_python\\Projetos_com_python\\Banking_System\\Banco_de_Dados.xlsx'
registar_client(WOORKBOOK)