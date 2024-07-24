from cadastro_de_cliente import Cliente
import os
from Validar_nif import validar_nif
import openpyxl
from openpyxl import workbook
import pandas as pd


class ColetarDados:

    def dados_do_cliente(self, arquivo_excel):
        print('Insira os seus dados:')
        
        """
        Collects data for a new client and creates a new instance of the Cliente class.
        
        This static method prompts the user to enter various personal identification and contact details for a new client. It then uses this information to create a new instance of the Cliente class. The collected data includes the client's name, age, gender, identification type, identification number, expiration date, marital status, NIF, residence details (country, district, street, postal code), birth details (date, country of birth), contact details (telephone number, email address), and account balance (minimum deposit of 100€).
        
        Returns:
            Cliente: A new instance of the Cliente class with the collected data.
        """
        self.nome = input('Nome: ')
        idade = input('Idade: ')
        sexo = input('Sexo [M/F]: ').upper()
        

        # validar Nome e NIF aqui mesmo
        while True:
            NIF = input('NIF: ')
            # validar NIF
            if validar_nif(NIF) == True: 
                break
            else:
                print(' NIF inválido! coloque um NIF valido')
                continue
        

        

        # Definir a função de comparação
        def _cliente_existe(name_cell, nif_cell):
            return name_cell == self.nome and (nif_cell) == int(NIF)
        
        def _nif_existe(nif_cell):
            return (nif_cell) == int(NIF)

        # validation of data      
        # Percorrer cada linha na planilha, começando na linha 2 para ignorar o cabeçalho
        try:
            # Carregar o arquivo Excel
            workbook = openpyxl.load_workbook(arquivo_excel)

            # Selecionar a planilha desejada
            sheet_selecionada = workbook.active
            for row in range(2, sheet_selecionada.max_row + 1): # Go through each line in excel
                self.name_cell = sheet_selecionada['A%s' % row].value  # percorre a toda linha da coluna "A - Nome"
                nif_cell = sheet_selecionada['H%s' % row].value   # percorre a toda linha da coluna "H - NIF" 

            # If the customer and the NIF exist, he says that the customer is already registered 
            if _nif_existe(nif_cell): # verify if NIF exists  
                if _cliente_existe(self.name_cell, nif_cell): # Verify is exists client associated with NIF
                    return None
        except FileNotFoundError:
            print()       

               
        # if not exist, it return False
        identificacao = input('Identificação [PC/T.R]: ')
        n_identificacao = input('Nº de Identificação: ')
        data_de_validade = input('Data de validade: ')
        estado_civil = input('Estado Civil: ')

        print('------------------------------------- ') 


        print('-------------Residência---------------') 
        pais = input('País: ')
        distrito = input('Distrito: ')
        rua = input('Rua: ')
        codigo_postal = input('Código-Postal: ')
        print('------------------------------------- ') 


        print('------------Naturalidade------------- ') 
        data_de_nasc = input('Data de Nascimento (dd/mm/aaaa):  ')
        # validar data de nascimento
        pais_de_nasc = input('País: ')
        print('------------------------------------- ') 


        print('--------------Contacto--------------- ') 
        telemovel = input('Telemóvel: ')
        e_mail = input('E-mail: ')
        print('------------------------------------- ') 
        
        os.system('cls') 
        print('------------Abrir Conta-------------- ')  
        
        print('Para concluir o processo, faça um depósito de no minimo 100€. \n Nota: Depois da abertura da conta o cliente pode poderá sacar o seu dinheiro quando quiser!!  ')
        saldo = float(input('Montante(€): '))
        

        # Criando um novo cliente
        novo_cliente = Cliente(self.nome,idade,sexo,identificacao,n_identificacao,data_de_validade,estado_civil, NIF,pais,distrito,rua, codigo_postal, data_de_nasc,pais_de_nasc,telemovel,e_mail, saldo)

        # cadastro = CadastrarCliente(nome,idade,sexo,identificacao,n_identificacao,data_de_validade,estado_civil, NIF,pais,distrito,rua, codigo_postal, data_de_nasc,pais_de_nasc,telemovel,e_mail, saldo)
        
        return novo_cliente
    

    #Registrar clientes permitindo a criação de uma senha
    def regitar_client(self, arquivo_excel):
        name_client = self.nome
        senha = input('Crie uma senha: ')

        try:
            # Carregar o arquivo Excel
            # Carregar o arquivo Excel
            """ workbook = openpyxl.load_workbook(arquivo_excel)

            # Verificar se a panilha "senhas_client" existe, caso contrario, criar uma nova
        
            if 'senhas_clint' not in workbook.sheetnames:
                workbook.create_sheet('senhas_client') """
            # Verificar se a panilha "senhas_client" existe, caso contrario, criar uma nova
        
            if os.path.exists(arquivo_excel):
                if 'senhas_client' in pd.ExcelFile(arquivo_excel).sheet_names:
                    df = pd.read_excel(arquivo_excel, sheet_name='senhas_client')
                else:
                    df = pd.DataFrame(columns=['Nome do Cliente', 'Senha'])
        

        

        except FileNotFoundError:
            df = pd.DataFrame(columns=['Nome do Cliente', 'Senha'])
            print('Arquivo não encontrado -> registar_client')
        

        novo_Cliente = {
                'Nome do Cliente': name_client,
                'Senha': senha
                
            }

        client_df = pd.DataFrame.from_dict([novo_Cliente])

        df = pd.concat([df, client_df], ignore_index=True)

        with pd.ExcelWriter(arquivo_excel, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='senhas_client', index=False)

        print( 'Senha guardada com Segurança -> registar_client')    
    
    
    

